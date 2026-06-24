"""Tests for the Excel case-log and the LaTeX report generator.

The .tex generation (no compile) and the Excel log are fast and run by default.
Actually compiling the PDF needs a LaTeX toolchain, so that check is opt-in.
"""
import os

import pytest

from turbine_design import StageInputs, DesignCoeffs, run_meanline, BladeChoices, design_blade
from turbine_design import caselog, latexreport


def _case():
    r = run_meanline(StageInputs(), DesignCoeffs())
    s2, s3, a = r["station2"], r["station3"], r["annulus"]
    bl = design_blade(s2["beta_deg"], s3["beta_deg"], s2["alpha_deg"], s3["alpha_deg"],
                      a["h2_m"], a["D_mean_m"], BladeChoices())
    return r, bl


# --- Excel case-log -------------------------------------------------------- #
def test_build_row_has_core_fields():
    r, bl = _case()
    row = caselog.build_row(r, blade=bl, case_name="ref")
    assert row["case"] == "ref"
    assert row["psi"] == 1.8 and row["n_blades"] == bl["n_blades"]
    assert row["all_pass"] is False           # reference violates M2 + h3/h2
    assert "M2" in row["violations"]


def test_append_creates_then_updates_in_place(tmp_path):
    log = str(tmp_path / "log.xlsx")
    r, bl = _case()
    caselog.append_case(log, r, blade=bl, case_name="A")
    caselog.append_case(log, r, blade=bl, case_name="B")
    import openpyxl
    ws = openpyxl.load_workbook(log).active
    assert ws.max_row == 3                     # header + 2 cases
    # re-running case A updates its row, not appends a new one
    caselog.append_case(log, r, blade=bl, case_name="A")
    ws = openpyxl.load_workbook(log).active
    assert ws.max_row == 3
    # a CFD-only update of case A fills the cfd columns without a new row
    caselog.append_case(log, cfd={"convergence": {"converged": True},
                                  "results": {"loading": {"suction_peak_Mis": 0.85},
                                              "exit_massaveraged": {"x": {"exit_mach": 0.69,
                                                                          "exit_flow_angle_deg": -70.0}}}},
                        case_name="A")
    ws = openpyxl.load_workbook(log).active
    assert ws.max_row == 3


# --- LaTeX generation (no compile) ----------------------------------------- #
def test_latex_tex_has_sections(tmp_path):
    r, bl = _case()
    out = latexreport.write_latex_report(r, str(tmp_path / "rep"), case_name="ref",
                                         blade=bl, compile_pdf=False)
    assert os.path.exists(out["tex"])
    tex = open(out["tex"]).read()
    assert r"\section{Mean-line design}" in tex
    assert r"\section{Rotor blade design}" in tex
    assert r"\begin{tabular}" in tex
    # intentional LaTeX math must survive escaping
    assert r"$\psi$" in tex or r"\psi" in tex


def test_latex_escaping_is_safe(tmp_path):
    """Free text with underscores must be escaped; cell math must not be."""
    r, _ = _case()
    r = dict(r); r["warnings"] = ["raise U_limit & check 50% margin"]
    out = latexreport.write_latex_report(r, str(tmp_path / "w"), case_name="a_b#c",
                                         compile_pdf=False)
    tex = open(out["tex"]).read()
    assert r"U\_limit" in tex and r"\&" in tex and r"\%" in tex


@pytest.mark.skipif(os.environ.get("RUN_LATEX_TESTS") != "1",
                    reason="set RUN_LATEX_TESTS=1 (needs a LaTeX toolchain) to compile")
def test_latex_compiles_to_pdf(tmp_path):
    r, bl = _case()
    out = latexreport.write_latex_report(r, str(tmp_path / "rep"), case_name="ref",
                                         blade=bl, compile_pdf=True)
    assert out["compiled"] and os.path.exists(out["pdf"])
