"""Accumulating Excel log of design test cases.

Every time a case is run, `append_case` adds one row (inputs, coefficients,
derived quantities, annulus, blade, constraints, and CFD results if present) to a
shared .xlsx workbook. Rows are keyed by `case` name: re-running the same case
name updates that row in place rather than duplicating it, so a design + its CFD
evaluation land on one line. PASS/FAIL is colour-coded.
"""
from __future__ import annotations
import os
from typing import Dict, Any, Optional, List

# (column key, header label, source)
COLUMNS = [
    ("case", "Case"), ("timestamp", "Timestamp"),
    ("P_shaft_MW", "P [MW]"), ("mdot", "mdot [kg/s]"), ("pi_C", "piC"),
    ("TIT_C", "TIT [C]"), ("n_stages", "stages"),
    ("psi", "psi"), ("phi", "phi"), ("DOR", "R"),
    ("U_mps", "U [m/s]"), ("specific_work_kJkg", "work [kJ/kg]"), ("P3_kPa", "P3 [kPa]"),
    ("M2", "M2"), ("M3", "M3"), ("beta2_deg", "b2 [deg]"), ("beta3_deg", "b3 [deg]"),
    ("alpha3_deg", "a3 [deg]"), ("turning_deg", "turn [deg]"), ("Mw3", "Mw3"),
    ("R_mean_m", "Rm [m]"), ("h2_mm", "h2 [mm]"), ("h3_mm", "h3 [mm]"),
    ("h3_h2", "h3/h2"), ("AN2", "AN2"),
    ("n_blades", "n_blades"), ("pitch_mm", "pitch [mm]"), ("chord_mm", "chord [mm]"),
    ("work_converged", "ml_converged"), ("U_capped", "U_capped"),
    ("all_pass", "constraints_pass"), ("violations", "violations"),
    ("cfd_converged", "cfd_converged"), ("cfd_exit_angle_deg", "cfd_a_exit [deg]"),
    ("cfd_exit_mach", "cfd_M_exit"), ("suction_peak_Mis", "Mis_peak_suc"),
]


def _round(v, n=4):
    try:
        return round(float(v), n)
    except Exception:
        return v


def build_row(result: Optional[Dict] = None, blade: Optional[Dict] = None,
              cfd: Optional[Dict] = None, case_name: str = "case",
              timestamp: str = "") -> Dict[str, Any]:
    row = {"case": case_name, "timestamp": timestamp}
    if result:
        d, a, c = result["derived"], result["annulus"], result["constraints"]
        co, inp = result["coefficients"], result["inputs"]
        s2, s3 = result["station2"], result["station3"]
        fe = result.get("feasibility", {})
        viol = [k for k, v in c.items() if isinstance(v, dict) and not v["pass"]]
        row.update({
            "P_shaft_MW": _round(inp["P_shaft"] / 1e6, 1), "mdot": _round(inp["mdot"], 1),
            "pi_C": _round(inp["pi_C"], 1), "TIT_C": _round(inp["TIT_C"], 0),
            "n_stages": inp["n_stages"],
            "psi": _round(co["psi"], 3), "phi": _round(co["phi"], 3), "DOR": _round(co["DOR"], 3),
            "U_mps": _round(d["U_mps"], 1), "specific_work_kJkg": _round(d["specific_work_Jkg"] / 1e3, 1),
            "P3_kPa": _round(d["P3_Pa"] / 1e3, 1),
            "M2": _round(s2["M"], 3), "M3": _round(s3["M"], 3),
            "beta2_deg": _round(s2["beta_deg"], 2), "beta3_deg": _round(s3["beta_deg"], 2),
            "alpha3_deg": _round(s3["alpha_deg"], 2), "turning_deg": _round(s3["turning_deg"], 1),
            "Mw3": _round(s3["Mrel"], 3),
            "R_mean_m": _round(a["R_mean_m"], 3), "h2_mm": _round(a["h2_m"] * 1e3, 1),
            "h3_mm": _round(a["h3_m"] * 1e3, 1), "h3_h2": _round(a["span_ratio_h3_h2"], 3),
            "AN2": _round(a["AN2_m2rpm2"], 0),
            "work_converged": fe.get("work_converged", True), "U_capped": fe.get("U_capped", False),
            "all_pass": c["all_pass"], "violations": ", ".join(viol) if viol else "",
        })
    if blade:
        row.update({"n_blades": blade["n_blades"], "pitch_mm": _round(blade["pitch_m"] * 1e3, 1),
                    "chord_mm": _round(blade["chord_m"] * 1e3, 1)})
    if cfd:
        conv = cfd.get("convergence", {}); res = cfd.get("results", {})
        ex = res.get("exit_massaveraged", {})
        ex0 = next(iter(ex.values()), {}) if ex else {}
        row.update({
            "cfd_converged": conv.get("converged"),
            "cfd_exit_angle_deg": _round(ex0.get("exit_flow_angle_deg"), 2) if ex0 else None,
            "cfd_exit_mach": _round(ex0.get("exit_mach"), 3) if ex0 else None,
            "suction_peak_Mis": _round(res.get("loading", {}).get("suction_peak_Mis"), 3),
        })
    return row


def append_case(log_path: str, result: Optional[Dict] = None, blade: Optional[Dict] = None,
                cfd: Optional[Dict] = None, case_name: str = "case",
                timestamp: str = "") -> str:
    """Append (or update by case name) one row in the .xlsx log. Returns the path."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    keys = [k for k, _ in COLUMNS]
    headers = [h for _, h in COLUMNS]
    new = build_row(result, blade, cfd, case_name, timestamp)

    if os.path.exists(log_path):
        wb = openpyxl.load_workbook(log_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "design cases"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

    # find an existing row with the same case name (col 1), else a new row
    target = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == case_name:
            target = r; break
    if target is None:
        target = ws.max_row + 1
        # carry forward nothing; blanks stay empty
    for ci, k in enumerate(keys, start=1):
        if k in new:
            ws.cell(target, ci, new[k])

    # colour the constraints_pass cell
    pass_col = keys.index("all_pass") + 1
    val = new.get("all_pass")
    if val is not None:
        fill = PatternFill("solid", fgColor=("C6EFCE" if val else "FFC7CE"))
        ws.cell(target, pass_col).fill = fill
    wb.save(log_path)
    return log_path
